import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl, time, os, fnmatch, csv, shutil, win32com.client
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains


def get_path_settings():
    with open('configs.csv', 'r') as csvfile:
        datareader = csv.reader(csvfile)

        for i, row in enumerate(datareader):
            if i == 0: continue
                
            return row[0], row[1], row[2]


def get_non_recorded_players(path, boss, boss_link_dict, nParses):

    chrome_options = Options()
    chrome_options.add_argument('load-extension=' + path)

    browser = webdriver.Chrome(chrome_options = chrome_options)
    time.sleep(5)

    browser.get(f"https://classic.warcraftlogs.com/zone/rankings/1011{boss_link_dict[boss]}&class=Druid&spec=Restoration&metric=hps")
    time.sleep(3)
    
    cookie = browser.find_element_by_class_name("cc-compliance")
    cookie.click()
    time.sleep(3)
    
    already_recorded_indices = get_latest_ranks(browser, boss, boss_link_dict, nParses)
    not_recorded = []
    
    browser.quit()
    time.sleep(1)
    
    for i in range(1, nParses):
        if i in already_recorded_indices: continue
        
        else: not_recorded.append(i)
            
    print(f"Ranks not in spreadsheet: {not_recorded}") 
    return not_recorded


def get_browser_page(boss, boss_link_dict, i):

    chrome_options = Options()
    chrome_options.add_argument('load-extension=' + r'C:\Users\Matth\Documents\1.42.4_6')

    browser = webdriver.Chrome(chrome_options = chrome_options)
    time.sleep(5)

    page = 1
    for j in range(1, i):
        if j % 100 == 0:
            page += 1
            
    if page == 1:
        boss_page_url = f'https://classic.warcraftlogs.com/zone/rankings/1011{boss_link_dict[boss]}&class=Druid&spec=Restoration&metric=hps'
        
    else:
        boss_page_url = f'https://classic.warcraftlogs.com/zone/rankings/1011{boss_link_dict[boss]}&class=Druid&spec=Restoration&metric=hps&page={page}'
    browser.get(boss_page_url)
    time.sleep(10)
    
    cookie = browser.find_element_by_class_name("cc-compliance")
    cookie.click()
    time.sleep(2)
    
    return browser


def get_boss_data_top_N_scraper(browser, i, boss, boss_link_dict):

    row = browser.find_element_by_id(f"row-{boss_link_dict[boss].split('=')[1]}-{i}")
    
    rank = row.text.split("\n")[0]
    name = row.text.split("\n")[1]
    date = row.text.split("\n")[-2].split(" ")[3] + " " + row.text.split("\n")[-2].split(" ")[4]
    server = row.text.split("\n")[-3].split(" ")[-2]
    region = row.text.split("\n")[-3].split(" ")[-1]
    HPS = row.text.split("\n")[-2].split(" ")[1].replace(",", "")
    duration = row.text.split("\n")[-2].split(" ")[-1]

    return rank, name, server, region, date, HPS, duration 


def get_latest_ranks(browser, boss, boss_link_dict, nParses):
    
    temp_df = pd.DataFrame(pd.np.empty((0, 3)))
    temp_df.columns = ["Rank", "Name", "Date"]
    
    page = 1

    for i in range(1, nParses + 50):
        if i%50 == 0: print(i)
        
        rank, name, server, region, date, HPS, duration = get_boss_data_top_N_scraper(browser, i, boss, boss_link_dict)
        if name in ['抄能力', '自然帅', 'Tables']: continue  # These players have broken reports, skip

        to_append = [rank, name, date]

        series = pd.Series(to_append, index = temp_df.columns)
        temp_df = temp_df.append(series, ignore_index = True)
        
        if i % 100 == 0: 
            page += 1
            browser.get(f'https://classic.warcraftlogs.com/zone/rankings/1011{boss_link_dict[boss]}&class=Druid&spec=Restoration&metric=hps&page={page}')
            time.sleep(10) 

    temp_df.to_csv(f"temp.csv", index = None, encoding='utf-8-sig')
    
    return update_ranks(boss)


def update_ranks(boss):

    temp = boss.replace(" ", "")

    wb = openpyxl.load_workbook(r"C:\Users\Matth\git\DataAnalysisWorkbooks\warcraftLogs\data\top_N_druids.xlsx")

    ws = wb.get_sheet_by_name(temp)
    sheet = wb[temp]
    rows = ws.max_row
    
    already_recorded_indices = []

    for row in ws.iter_rows():
        if row[0].value == "Rank": continue  # Skip header

        with open(f'temp.csv', 'r', encoding='utf-8-sig') as csvfile:

            for i, line in enumerate(csvfile.readlines()):

                line = line[ : -1]  # Remove the newline character at the end of each string
                elements = line.split(",")      
                
                if [elements[1], elements[2]] == [row[1].value, row[3].value] and int(float(elements[0])) != row[0].value:
                    print(f"Rank updated: {row[0].value} to {int(float(elements[0]))}, {elements[1]}, {elements[2]}")
                    row[0].value = int(float(elements[0]))
                    
                # If character already in spreadsheet (could be correct rank or not), don't want to rescrape that character.
                if [elements[1], elements[2]] == [row[1].value, row[3].value]:
                    already_recorded_indices.append(row[0].value)
                    
    wb.save(r"C:\Users\Matth\git\DataAnalysisWorkbooks\warcraftLogs\data\top_N_druids.xlsx")
    wb.save(r"C:\Users\Matth\git\DataAnalysisWorkbooks\warcraftLogs\data\top_N_druids_backup.xlsx")
    os.remove(f"temp.csv")
        
    return already_recorded_indices


def get_nHealers(browser):
    row = browser.find_elements_by_class_name("composition-row")
    
    return len(row[2].text.split("\n")) - 1
    

def get_tanks(browser):
    click_on_element_by_id(browser, "filter-summary-tab")  
    time.sleep(1)
    
    row = browser.find_elements_by_class_name("composition-row")
    
    tanks = row[0].text.split("\n")
    tanks.pop(0)
    
    return tanks


def click_on_element_by_id(browser, id_tag):    
    tag = browser.find_element_by_id(id_tag)
    time.sleep(1)
    
    for j in range(20):
        action = ActionChains(browser)
        action.move_to_element(tag)
        time.sleep(0.5)
        
        try:
            action.perform()
            time.sleep(0.5)
            action.click()
            action.perform()
            break
        
        except:
            body = browser.find_element_by_css_selector('body')
            body.click()
            body.send_keys(Keys.PAGE_UP) 
            
            
def get_spell_info(browser, total_HPS):

    # Lifebloom (tick), Lifebloom (bloom), Rejuv, Regrowth, Swiftmend
    spell_percent_HPS_dict = {"ability-33763-0" : 0, "ability-33778-0" : 0, "ability-26982-0" : 0, "ability-26980-0" : 0, "ability-18562-0" : 0}
    spell_ids = ['ability-33763-0', 'ability-33778-0', 'ability-26982-0', 'ability-26980-0', 'ability-18562-0']
    LB_uptime = 0
    
    for spell in spell_ids:
        
        try:
            search = browser.find_element_by_id(spell)

            td_p_input = search.find_element_by_xpath('.//ancestor::td')
            tr_p_input = search.find_element_by_xpath('.//ancestor::tr')
            td_p_input = tr_p_input.find_elements(By.XPATH, 'td')
            
            spell_percent_HPS_dict[spell] = round(float(td_p_input[9].text.replace(",", "")) / float(total_HPS), 2)
            
            if spell == 'ability-33763-0':
                LB_uptime = td_p_input[7].text
            
        except:
            pass
        
    return spell_percent_HPS_dict['ability-33763-0'], spell_percent_HPS_dict['ability-33778-0'], spell_percent_HPS_dict['ability-26982-0'], spell_percent_HPS_dict['ability-26980-0'], spell_percent_HPS_dict['ability-18562-0'], LB_uptime


def check_spriest(browser):
    body = browser.find_element_by_css_selector('body')
    body.send_keys(Keys.PAGE_UP)
    time.sleep(1)
    
    click_on_element_by_id(browser, "filter-resources-tab") 
    time.sleep(3)
    
    click_on_element_by_id(browser, "filter-resource-selection-dropdown")
    time.sleep(3)
    
    a = browser.find_element_by_id("spell-100")
    time.sleep(1)
    a.click()
    
    time.sleep(2)
    b = browser.find_elements_by_class_name("main-table-name")
    
    # German Spriest?
    for i in range(len(b)):
        if b[i].text in ['Vampiric Touch', '吸血鬼之触', '흡혈의 손길', 'Прикосновение вампира', 'Toucher vampirique']:
            return 'Yes'
            
    return 'No'


def check_buffs(browser):
    innervate, bloodlust, powerInfusion, naturesGrace = 'No', 'No', 'No', 'No'
    
    body = browser.find_element_by_css_selector('body')
    body.send_keys(Keys.PAGE_UP)
    time.sleep(1)
    
    click_on_element_by_id(browser, "filter-buffs-tab")
    time.sleep(2)
    
    try:
        a = browser.find_element_by_id("main-table-256-0")
        b = a.text.split('\n')
        if len(fnmatch.filter(b, 'Innervate??')) > 0 or len(fnmatch.filter(b, '激活??')) > 0  or len(fnmatch.filter(b, '정신 자극??')) > 0 or len(fnmatch.filter(b, 'Озарение??')) > 0 or len(fnmatch.filter(b, 'Innervation??')) > 0 or len(fnmatch.filter(b, 'Anregen??')) > 0: 
            innervate = 'Yes'
       
    except: 
        pass
               
    try:
        a = browser.find_element_by_id("main-table-1-0")
        b = a.text.split('\n')
               
        if len(fnmatch.filter(b, 'Bloodlust??')) > 0 or len(fnmatch.filter(b, 'Heroism??')) > 0:
            bloodlust = 'Yes'
            
        elif len(fnmatch.filter(b, '嗜血??')) > 0 or len(fnmatch.filter(b, '英勇??')) > 0:
            bloodlust = 'Yes'
            
        elif len(fnmatch.filter(b, '피의 욕망??')) > 0 or len(fnmatch.filter(b, '영웅심??')) > 0: 
            bloodlust = 'Yes'
            
        elif len(fnmatch.filter(b, 'Кровожадность')) > 0 or len(fnmatch.filter(b, 'Героизм??')) > 0:
            bloodlust = 'Yes'
            
        elif len(fnmatch.filter(b, 'Furie sanguinaire??')) > 0 or len(fnmatch.filter(b, 'Héroïsme??')) > 0:
             bloodlust = 'Yes'
                
        # German bloodlust?
        #elif len(fnmatch.filter(b, '')) > 0 or len(fnmatch.filter(b, 'Heldentum??')) > 0:
        elif len(fnmatch.filter(b, 'Heldentum??')) > 0:
            bloodlust = 'Yes'
            
        # German PI?
        if len(fnmatch.filter(b, 'Power Infusion??')) > 0 or len(fnmatch.filter(b, '能量灌注??')) > 0 or len(fnmatch.filter(b, '마력 주입??')) > 0 or len(fnmatch.filter(b, 'Придание сил??')) > 0 or len(fnmatch.filter(b, 'Infusion de puissance??')) > 0: 
            powerInfusion = 'Yes'
                     
        if len(fnmatch.filter(b, "Nature's Grace???")) > 0 or len(fnmatch.filter(b, '自然之赐???')) > 0  or len(fnmatch.filter(b, '자연의 은혜???')) > 0 or len(fnmatch.filter(b, 'Благоволение природы???')) > 0 or len(fnmatch.filter(b, 'Grâce de la nature???')) > 0 or len(fnmatch.filter(b, 'Anmut der Natur???')) > 0:
            naturesGrace = 'Yes'
            
    except:
        pass
    
    return innervate, bloodlust, powerInfusion, naturesGrace


# When running scrapers in parallel the cast CSVs will tend to download at the same time. Unfortunately these download
# with the same name. To find the specific player's file, open the csvs and see whose file it is.
def check_if_player_file(name, filename):
    with open(filename, 'r', encoding='utf-8-sig') as csvfile:
        for i, line in enumerate(csvfile.readlines()):
            if i == 0: continue  # Skip header

            line = line[ : -1]  # Remove the newline character at the end of each string
            elements = line.split(",")

            a = elements[3].split(" ")
            b = a[0].replace('"', "")
            
            if name == b:
                return True
            
            else:
                return False
            
            
def download_csv(browser, temp_url, id_tag, download_path, path, name, nCores):
    body = browser.find_element_by_css_selector('body')
    body.send_keys(Keys.PAGE_UP)
    time.sleep(1)
    
    click_on_element_by_id(browser, "filter-events-tab")    
    time.sleep(1)

    click_on_element_by_id(browser, id_tag)
    time.sleep(1)
    
    body.send_keys(Keys.END)
    time.sleep(1)
    tag = browser.find_element_by_class_name("buttons-csv")
    tag.click()
    time.sleep(3)
    
    for i in range(nCores):
        
        if i==0:
            try:
                if check_if_player_file(name, "C:/Users/Matth/Downloads/Warcraft Logs - Combat Analysis for Warcraft.csv"):
                    #print(f"{name} file is Warcraft Logs - Combat Analysis for Warcraft.csv")
                    filename = "C:/Users/Matth/Downloads/Warcraft Logs - Combat Analysis for Warcraft.csv"
                    break
                    
                else: continue
                
            except: pass
                
        else:
            try: 
                if check_if_player_file(name, f"C:/Users/Matth/Downloads/Warcraft Logs - Combat Analysis for Warcraft ({i}).csv"):
                    #print(f"{name} file is Warcraft Logs - Combat Analysis for Warcraft ({i}).csv")
                    filename = f"C:/Users/Matth/Downloads/Warcraft Logs - Combat Analysis for Warcraft ({i}).csv"
                    break
                    
                else: continue
                
            except: pass
            
    shutil.move(filename, path)
    time.sleep(1)
    
    
def clean_cast_sequence_csv(name, path_to_data_dir):
    correct_csv_whitespace(path_to_data_dir + f"\csv\cast_sequence_{name}")
    
    df = pd.read_csv(path_to_data_dir + f"\csv\cast_sequence_{name}.csv", encoding='utf-8-sig')
    
    df = df.drop(['Unnamed: 4'], axis=1)
    
    temp_df = df['Time'].str.split(":", expand=True)
    df['Minute'] = temp_df[0]
    df['Second'] = temp_df[1]

    df['Minute'] = pd.to_numeric(df['Minute'])
    df['Second'] = pd.to_numeric(df['Second'])

    temp_df = df['Ability'].str.split(" ", expand=True)
    try:
        df['Ability'] = temp_df[0]
        df['Cast Time'] = temp_df[1]
        
    except:  # Pandas will crash if the Druid did not cast any abilities with a cast time
        df['Ability'] = temp_df[0] 

    temp_df = df['Source → Target'].str.split(" ", expand=True)
    
    # If DBM puts a marker on a tank it can mess up the pandas dataframe. Check and fix if this happens
    fixed_target = []
    for i, row in (temp_df.iterrows()):
        if row[2] != "": 
            fixed_target.append(row[2])
            
        else:
            fixed_target.append(row[3])
            
    df['Target'] = fixed_target

    df['Boss Target'] = np.nan

    df = df.drop(['Source → Target'], axis=1)
    
    return df


# Fix orrcurances where there is a double whitespace in the pandas dataframe
def correct_csv_whitespace(path):

    filename = path+'.csv'
    with open(filename, 'r', encoding='utf-8-sig') as csvfile:
        outputFile = open(path+"_fixed.csv", "w", encoding='utf-8-sig')
        
        datareader = csv.reader(csvfile)
        outputWriter = csv.writer(outputFile, lineterminator='\n')

        for row in datareader:
            if len(row) == 0: continue

            for i, item in enumerate(row):
                row[i] = row[i].replace("  ", " ")
                row[i] = row[i].replace('"', '')
            outputWriter.writerow(row)
        outputFile.close()

    shutil.move(path + "_fixed.csv", path + ".csv")   
    
    
def fix_cast_time(df):
    temp = []
    for i, row in df.iterrows():
        if row['Ability'] in ['Regrowth', 'Rebirth', '癒合', '愈合', '재생', '환생', 'Восстановление', 'Rétablissement', 'Renaissance', 'Nachwachsen', 'Wiedergeburt'] and row["Cast Time"] is None: 
            temp.append(row["Time"])
        
        elif row['Ability'] in ['Regrowth', 'Rebirth', '癒合', '愈合', '재생', '환생', 'Восстановление', 'Rétablissement', 'Renaissance', 'Nachwachsen', 'Wiedergeburt'] and row["Cast Time"] != "Canceled":
            a = str(datetime.strptime(str(row['Minute']) + ":" + str(row['Second']), "%M:%S.%f") - datetime.strptime(row['Cast Time'], "%S.%f"))
            a = a.split(".")
            if len(a) == 1: a.append('000000')
            b = a[0].split(":")
            c = b[1] + ":" + b[2] + "." + a[1][:3]
            temp.append(c)

        else:
            temp.append(row["Time"])

    df["Time"] = temp
    
    temp_df = df['Time'].str.split(":", expand=True)
    df['Second'] = temp_df[1]
    df['Second'] = pd.to_numeric(df['Second'])
    
    return df


# Rotation calculator. This function checks when lifebloom is refreshed on the primary tank and what sequence of spells (Instant, Regrowth)
# are cast before the initial lifebloom is refreshed. The rotations are tallied in a rotation dictionary and a final count is returned.
def calculate_rotations(df, boss, boss_tanks, LB_uptime, verbose, verbose_rotation):

    GCD_time = {"t1" : 0, "t2" : 0}
    rotations_dict = {}

    for i in range(3):
        if len(boss_tanks) < 3: boss_tanks.append("X")

    sequence, LB_tank_flags = [], [False, False, False]
    starting_tank, LB_on_tank = None, False
    
    if verbose_rotation:
        print("Printing full rotation..")
        print("--------")

    for i, row in (df.iterrows()):
        
        # Ignore p1 for ROS
        if boss == "Reliquary of Souls" and row["Ability"] in ["Wrath", "Moonfire", "Starfire"]: continue
        
        # If seq gets to 4 casts, record and reset. TODO: For hasted rotations this can be 5
        if len(sequence) == 5:
            if verbose_rotation: print(sequence)
            rotations_dict = count_rotations(sequence, rotations_dict)
            LB_tank_flags = [False, False, False]
            sequence = []

        # Ignore casts that are off the global cooldown
        # Spells that should definitely be off GCD (trinkets, mana potion, dark rune):
        if row["Ability"] in ["Hopped", "Essence", "Dark", "Restore",
                              "恢复法力", "黑暗符文", "自然迅捷", "殉难者精华", "佳酿",
                              "마나", "순교자의", "홉", "생명의",
                              "Сущность", "Приподнятый", "Восполнение",
                              "Houblonné", "Restauration", "Rune",
                              "Essenz", "Hopfen", "Mana"]: continue
        # Check GCD manually to catch any not included in the above:
        temp = delta_t(row, GCD_time)
        GCD_time = temp[0]
        if not temp[1]: 
            continue  
            
        # If LB refreshed on primary tank, record and restart the sequence
        if row['Ability'] in ['Lifebloom', '生命绽放', '피어나는', 'Жизнецвет', 'Fleur', 'Blühendes'] and row["Target"] == starting_tank:
            if len(sequence) > 0:
                if verbose_rotation: print(sequence)
                rotations_dict = count_rotations(sequence, rotations_dict)

            sequence = []
            LB_tank_flags = [False, False, False]
            for j in range(3):
                if row["Target"] == boss_tanks[j]: LB_tank_flags[j] = True

            sequence.append('Lifebloom')
            continue

        # Otherwise check if LB placed on a tank, restart sequence if necessary
        for j in range(3):

            if row['Ability'] in ['Lifebloom', '生命绽放', '피어나는', 'Жизнецвет', 'Fleur', 'Blühendes'] and row["Target"] == boss_tanks[j] and not LB_tank_flags[j]:
                
                if True not in LB_tank_flags:
                    starting_tank = boss_tanks[j]
                    if len(sequence) > 0:
                        if verbose_rotation: print(sequence)
                        rotations_dict = count_rotations(sequence, rotations_dict)
                    sequence = []

                sequence.append('Lifebloom')
                LB_tank_flags[j] = True
                LB_on_tank = True
                
        if LB_on_tank:
            LB_on_tank = False
            continue


        # Track regrowth casts specifically
        if row['Ability'] in ["Regrowth", '癒合', '愈合', '재생', 'Восстановление', 'Rétablissement']:
            sequence.append("Regrowth")

        # All other casts are considered 'instant'. (Including spells with cast times like Rebirth or Drums)
        else: 
            sequence.append("Instant")     
            
    if verbose_rotation: print("--------")

    return update_rotation_dict(rotations_dict, LB_uptime, verbose)


def update_rotation_dict(rotations_dict, LB_uptime, verbose):
    
    # Convert rotation counts to percentages
    total = 0
    for key in rotations_dict:
        total += rotations_dict[key]

    for key in rotations_dict:
        rotations_dict[key] = round(float(rotations_dict[key]) / total, 3)
        
    nontank_rotation_percent = 0
    if verbose: print(f"Lifebloom uptime %: {LB_uptime}")
    for key in rotations_dict:
        if verbose: print(key, rotations_dict[key])
        if key[0] == '0':
            nontank_rotation_percent += rotations_dict[key]
    if verbose: print(f"Non-tank rotations %: {round(nontank_rotation_percent, 2)}")
    
    # Check if player is actually rotating on the tank and not just raid healing.
    if nontank_rotation_percent >= 0.5:# and float(str(LB_uptime).replace("%", "")) < 50.0: 
        rotating_on_tank = "No"
        
    else:
        rotating_on_tank = "Yes"
        
    if verbose: print(f"Rotating on tank?: {rotating_on_tank}")
    
    # Pick out the top two rotations used
    max_key = max(rotations_dict, key = rotations_dict. get)
    rotation1 = max_key
    rotation1_percent = rotations_dict[max_key]

    rotations_dict.pop(max_key)

    try:
        max_key = max(rotations_dict, key=rotations_dict. get)
        rotation2 = max_key
        rotation2_percent = rotations_dict[max_key]
        
    except ValueError:
        rotation2 = 'None'
        rotation2_percent = 0.0

    return rotation1, rotation1_percent, rotation2, rotation2_percent, rotating_on_tank
    
    
def countX(lst, x):
    return lst.count(x)


def count_rotations(lst, rotations_dict):
    
    # Ignore sequences length 2 or less
    if len(lst) <= 2: 
        return rotations_dict
    
    else:
        LB_count = countX(lst, "Lifebloom")
        I_count = countX(lst, "Instant")
        RG_count = countX(lst, "Regrowth")
    
        # Add rotation key to dictionary if it isn't a key yet       
        rotations_dict.setdefault(f'{LB_count}LB {I_count}I {RG_count}RG', 0)
        rotations_dict[f'{LB_count}LB {I_count}I {RG_count}RG'] += 1

        return rotations_dict

        
def delta_t(row, time):  # Spell casts that are off-GCD (trinkets, etc) should not go into the rotation
    time["t1"] = time["t2"]
    time["t2"] = datetime.strptime(str(row['Minute']) + ":" + str(row['Second']), "%M:%S.%f") 
    
    if type(time["t2"]) == datetime and type(time["t1"]) == datetime:
        
        try:
            if float(str((time["t2"] - time["t1"]))[5:]) < 0.4:
                return [time, False]
        except ValueError:
            return [time, True]

    return [time, True]


def export_to_csv(path_to_data_dir, boss, to_append, player_df, char_name):#, filename, convertRank):
    series = pd.Series(to_append, index = player_df.columns)
    player_df = player_df.append(series, ignore_index = True)

    # Export dataframe to csv
    player_df.to_csv(path_to_data_dir + f"\csv\{boss.replace(' ', '')}_{char_name}.csv", index = None, encoding='utf-8-sig')
    
    
def add_rows_to_xlsx(path_to_data_dir, boss, filename, convertRank):
    temp = boss.replace(" ", "")

    path_to_data_dir = r'C:\Users\Matth\git\DataAnalysisWorkbooks\warcraftLogs\data'

    dirs = os.listdir(path_to_data_dir + f"\csv")

    for file in dirs:
        wb = openpyxl.load_workbook(path_to_data_dir + f"\{filename}.xlsx")

        ws = wb.get_sheet_by_name(temp)
        sheet = wb[temp]
        rows = ws.max_row

        if file[0:3] != "cast":
            print(file)

            with open(path_to_data_dir + f"\csv\{file}", 'r', encoding='utf-8-sig') as csvfile:

                for i, line in enumerate(csvfile.readlines()):
                    if i == 0: continue  # Skip header

                    line = line[ : -1]  # Remove the newline character at the end of each string
                    elements = line.split(",")

                    for i2, element in enumerate(elements):
                        if i2 == 0 and convertRank: 
                            sheet[chr(i2 + 65) + str(i + rows)] = int(float(element))
                        else: 
                            sheet[chr(i2 + 65) + str(i + rows)] = element

        wb.save(path_to_data_dir + f"\{filename}.xlsx")
        wb.save(path_to_data_dir + f"\{filename}_backup.xlsx")

        

def sort_excel(path_to_data_dir, boss, filename):
    temp = boss.replace(" ", "")
    

    wb = openpyxl.load_workbook(path_to_data_dir + f"\{filename}.xlsx")
    ws = wb.get_sheet_by_name(temp)
    sheet = wb[temp]
    rows = ws.max_row

    if filename == 'character_data':
        order_cell, ordering = 'E', 2

    else:
        order_cell, ordering = 'A', 1

    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(path_to_data_dir + f"\{filename}.xlsx")
    ws = wb.Worksheets(temp)
    ws.Range('A2:W'+str(rows+1)).Sort(Key1 = ws.Range(f'{order_cell}1'), Order1 = ordering, Orientation = 1)
    wb.Save()
    excel.Application.Quit()
    

def remove_xlsx_duplicates(path_to_data_dir, boss, filename):

    print("Removing spreadsheet duplicates..")

    while True:

        temp = boss.replace(" ", "")
        
        wb = openpyxl.load_workbook(path_to_data_dir + f"\{filename}.xlsx")

        ws = wb.get_sheet_by_name(temp)
        sheet = wb[temp]
        rows = ws.max_row

        temp = ["", ""]
        duplicate_count = 0

        for i, row in enumerate(ws.iter_rows()):

            temp[0] = temp[1]
            temp[1] = row[0].value

            if i==0: continue

            if temp[0] == temp[1]:
                duplicate_count += 1
                ws.delete_rows(i)

        wb.save(path_to_data_dir + f"\{filename}.xlsx")
        wb.save(path_to_data_dir + f"\{filename}_backup.xlsx")

        if duplicate_count == 0:
            print("Duplicate removal complete!")
            break
    
    
def clean_csv_dir(path_to_data_dir):
    dirs = os.listdir(path_to_data_dir + rf"\csv")

    for file in dirs:
        os.remove(path_to_data_dir + rf"\csv\{file}")